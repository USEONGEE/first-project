from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook('C:/Users/shdbt/Desktop/22년도 학술제/부산_중구_노인복지시설.xlsx',data_only = True)
ws = wb['부산_중구_노인복지시설']

w = Workbook()
w.__init__('C:/Users/shdbt/Desktop/22년도 학술제/부산_중구_노인복지시설.xlsx',data_only = True)
w.sheetnames('부산_중구_노인복지시설')


row = 2
column1 = 5
column2 = 6
e = []    
f = []
while(ws.cell(row,column1).value != None) :
    e.append(ws.cell(row,column1).value)
    f.append(ws.cell(row,column2).value)
    row += 1


all = []
for i in range(len(e)) :
    all.append('부산광역시 중구 ' + e[i] + ' ' + f[i])
print(type(all[0]))

for i in range(len(e)) :
    ws.cell(i+1,10).value = all[i]
    
wb.save('C:/Users/shdbt/Desktop/22년도 학술제/부산_중구_노인복지시설.xlsx')

